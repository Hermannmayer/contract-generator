"""PI / COMMERCIAL INVOICE Excel generator — standard trade format."""

import io, os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.worksheet.properties import PageSetupProperties

def _set_cell(ws, r, c, val, font=None, align=None, border=None):
    cell = ws.cell(row=r, column=c, value=val)
    if font: cell.font = font
    if align: cell.alignment = align
    if border: cell.border = border
    return cell

def _merge_and_border(ws, r1, c1, r2, c2):
    """Merge range and apply thin_border to the merged area."""
    for r in range(r1, r2+1):
        for c in range(c1, c2+1):
            ws.cell(row=r, column=c).border = thin_border_ref
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)

# Border singletons (must be defined after function)
thin_border_ref = Border(left=Side(style="thin", color="000000"), right=Side(style="thin", color="000000"),
                         top=Side(style="thin", color="000000"), bottom=Side(style="thin", color="000000"))

def _merge_carton_groups(ws, products, start_row, end_row, columns):
    groups = {}
    for i, prod in enumerate(products):
        gid = prod.get("carton_group", 0)
        if gid:
            rn = start_row + i
            if gid not in groups: groups[gid] = {"start": rn, "end": rn}
            else: groups[gid]["end"] = rn
    for gid, rng in groups.items():
        if rng["start"] == rng["end"]: continue
        for col in columns:
            cl = col + 64
            if cl > 90: cl += 1  # handle K
            ws.merge_cells(f"{chr(cl)}{rng['start']}:{chr(cl)}{rng['end']}")

def generate_pi_excel(data: dict, image_paths: dict = None, seal_path: str = None) -> bytes:
    if image_paths is None: image_paths = {}
    wb = Workbook(); ws = wb.active; ws.title = "COMMERCIAL INVOICE"

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center")
    hdr_f = Font(name="Arial", size=9, bold=True); sub_f = Font(name="Arial", size=8, bold=True)
    title_f = Font(name="Arial", size=14, bold=True); en_f = Font(name="Arial", size=11, bold=True)
    addr_f = Font(name="Arial", size=10); norm_f = Font(name="Arial", size=11)
    small_f = Font(name="Arial", size=10); marks_f = Font(name="Arial", size=9, bold=True)
    big_b = Font(name="Arial", size=12, bold=True); bold_f = Font(name="Arial", size=10, bold=True)
    sep_b = Border(top=Side("thin","000000"), bottom=Side("thin","000000"),
                   left=Side("thin","000000"), right=Side("thin","000000"))

    for cl, w in [("A",18),("B",20),("C",12),("D",8),("E",12),("F",9),("G",9),("H",16),("I",13),("J",13),("K",14)]:
        ws.column_dimensions[cl].width = w

    client = data.get("client", {}); port_from = data.get("port_from", {}); port_to = data.get("port_to", {})
    products = data.get("products", []); deposit_rate = data.get("deposit_rate", "30%")
    port_name = port_from.get("name_en", "QINGDAO")
    seller_en = "YOUR COMPANY NAME"
    seller_addr = "YOUR COMPANY ADDRESS"
    china_name = "贵公司名称"
    MAXC = 11  # A-K

    row = 1

    # ══════ R1-4: Header (centered, merged A-K) ══════
    for val, ft in [(china_name, addr_f), (seller_en, en_f), (seller_addr, addr_f), ("COMMERCIAL INVOICE", title_f)]:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=MAXC)
        _set_cell(ws, row, 1, val, ft, center)
        row += 1

    # ══════ R5-7: PO/Date/Supplier — label I(9), value J(10) ══════
    for lbl, val in [("PO No. :", data.get("contract_no","")),
                     ("Date:", data.get("date","")),
                     ("Supplier Code:", data.get("supplier_code",""))]:
        _set_cell(ws, row, 9, lbl, norm_f, left)
        _set_cell(ws, row, 10, val, norm_f, left)
        row += 1

    # ══════ R8-10: BUYER ══════
    for lbl, val in [("BUYER", client.get("name","")),
                     ("ADDRESS", client.get("address","")),
                     ("TEL", client.get("phone","") or "")]:
        _set_cell(ws, row, 1, lbl, norm_f, left)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
        _set_cell(ws, row, 2, val, norm_f, left)
        row += 1

    row += 1  # R11 spacer

    # ══════ R12-13: SELLER ══════
    for lbl, val in [("SELLER", seller_en), ("ADDRESS", seller_addr)]:
        _set_cell(ws, row, 1, lbl, norm_f, left)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
        _set_cell(ws, row, 2, val, norm_f, left)
        row += 1
    row += 1  # R14 spacer

    # ══════ R15: Transport (no wrap, compact height) ══════
    l_nw = Alignment(horizontal="left", vertical="center"); c_nw = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    _set_cell(ws, row, 2, f"From: {port_name},CHINA", norm_f, l_nw)
    _set_cell(ws, row, 6, f"To: {port_to.get('name_en','')},JAPAN", norm_f, c_nw)
    _set_cell(ws, row, 9, "BY SEA", norm_f, c_nw)
    ws.row_dimensions[row].height = 18
    row += 1

    # Divider line (A-J only, not K)
    for c in range(1, 11):
        ws.cell(row=row, column=c).border = Border(bottom=Side("thin","000000"))
    row += 1

    # ══════ Product table header (2-row) — borders BEFORE merge ══════
    h1 = row; h2 = h1 + 1
    for c in range(1, 11):
        ws.cell(row=h1, column=c).border = thin_border_ref
        ws.cell(row=h2, column=c).border = thin_border_ref

    h1_defs = [(1,"Marks & Numbers",1,2),(2,"Description of Goods",3,2),
               (4,"color",4,2),(5,"SIZE",5,2),(6,"PCS",6,2),(7,"PCS",7,2),
               (8,"CARTON",8,2),(9,f"FOB {port_name}",9,1),(10,"Amount",10,1)]
    for col, val, ec, rs in h1_defs:
        ws.merge_cells(start_row=h1, start_column=col, end_row=h1+rs-1, end_column=ec)
        bd = sep_b if col in (9,10) else thin_border_ref
        _set_cell(ws, h1, col, val, hdr_f, center, bd)

    for col in [9, 10]:
        _set_cell(ws, h2, col, "USD", sub_f, center, sep_b)
    row = h2 + 1

    # ══════ Product data ══════
    prod_start = row
    if products:
        # Apply borders BEFORE merge for Marks column
        for mr in range(row, row + len(products)):
            ws.cell(row=mr, column=1).border = thin_border_ref
        ws.merge_cells(start_row=row, start_column=1, end_row=row+len(products)-1, end_column=1)
        _set_cell(ws, row, 1, "3COINS\nCTN NO.\nMADE IN CHINA", marks_f, center)

    for idx, prod in enumerate(products):
        _set_cell(ws, row, 2, prod.get("item_no",""), small_f, center, thin_border_ref)
        _set_cell(ws, row, 3, None, None, None, thin_border_ref)  # image placeholder
        _set_cell(ws, row, 4, prod.get("color",""), small_f, center, thin_border_ref)
        _set_cell(ws, row, 5, prod.get("size",""), small_f, center, thin_border_ref)
        _set_cell(ws, row, 6, prod.get("qty",""), small_f, center, thin_border_ref)
        v7 = prod.get("pcs_per_carton","") if prod.get("carton_group_first",True) else ""
        _set_cell(ws, row, 7, v7, small_f, center, thin_border_ref)
        v8 = prod.get("cartons","") if prod.get("carton_group_first",True) else ""
        _set_cell(ws, row, 8, v8, small_f, center, thin_border_ref)
        _set_cell(ws, row, 9, prod.get("unit_price_usd",""), small_f, center, thin_border_ref)
        _set_cell(ws, row, 10, prod.get("amount_usd",""), small_f, center, thin_border_ref)

        # Embed product image (openpyxl anchors to top-left of cell)
        img_key = str(idx)
        if img_key in image_paths and os.path.exists(image_paths[img_key]):
            try:
                from openpyxl.drawing.image import Image as OpxImage
                img = OpxImage(image_paths[img_key]); img.width = 55; img.height = 55
                ws.add_image(img, f"C{row}")
                ws.row_dimensions[row].height = 60
            except Exception:
                pass
        ws.row_dimensions[row].height = max(ws.row_dimensions[row].height or 20, 20)
        row += 1

    _merge_carton_groups(ws, products, prod_start, row-1, [7,8])

    # ══════ TOTAL ══════
    total_amt = data.get("total_amount_usd","0")
    total_val = float(str(total_amt).replace(",",""))
    ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=5)
    _set_cell(ws, row, 4, "TOTAL :", Font(name="Arial", size=10), Alignment(horizontal="right", vertical="center"))
    _set_cell(ws, row, 6, data.get("total_qty",""), Font(name="Arial", size=9), center)
    _set_cell(ws, row, 8, data.get("total_cartons",""), Font(name="Arial", size=9), center)
    _set_cell(ws, row, 10, f"${total_amt}", big_b, center)
    row += 1

    # ══════ Deposit — label I(9), amount J(10) ══════
    if total_val >= 10000:
        try:
            dep_pct = float(deposit_rate.replace("%",""))/100
            dep_amt = round(total_val*dep_pct,2)
        except: dep_amt = 0
        _set_cell(ws, row, 9, "30% DEPOSIT:", bold_f, left)
        _set_cell(ws, row, 10, f"${dep_amt:,.2f}", bold_f, left)
        row += 1
    row += 1

    # ══════ PAYMENT + ETD ══════
    if total_val >= 10000:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=11)
        _set_cell(ws, row, 1, f"PAYMENT TERM : {data.get('payment','')}", norm_f, left); row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=11)
        _set_cell(ws, row, 1, f"ETD : {data.get('etd','')}", norm_f, left); row += 2
    else:
        row += 1

    # ══════ Signatures + Seal ══════
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    _set_cell(ws, row, 1, "BUYER : ________________________", norm_f)
    ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=11)
    _set_cell(ws, row, 6, f"SELLER : {seller_en}", norm_f)
    # Place seal at column H over the text, no row height adjustment
    if seal_path and os.path.exists(seal_path):
        try:
            from openpyxl.drawing.image import Image as OpxImage
            seal = OpxImage(seal_path); seal.width = 150; seal.height = 150
            ws.add_image(seal, f"H{row}")
        except: pass

    ws.page_setup.orientation = "portrait"; ws.page_setup.paperSize = 9
    ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_margins.left = 0.3; ws.page_margins.right = 0.3
    ws.page_margins.top = 0.4; ws.page_margins.bottom = 0.4
    out = io.BytesIO(); wb.save(out); out.seek(0)
    return out.getvalue()

def generate_pi_excel_with_images(data: dict, image_paths: dict = None, seal_path: str = None) -> bytes:
    return generate_pi_excel(data, image_paths, seal_path)
