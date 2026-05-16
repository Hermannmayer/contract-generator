"""Factory contract (成品定做合同) Excel generator."""

import io
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image as OpxImage
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties


def num2cn(amount):
    """Convert numeric amount to Chinese uppercase (e.g. 46970 -> 肆万陆仟玖佰柒拾元整)."""
    digits = "零壹贰叁肆伍陆柒捌玖"
    units = ["仟", "佰", "拾", ""]
    big_units = ["", "万", "亿"]
    try:
        amt = float(str(amount).replace(",", ""))
    except (ValueError, TypeError):
        return "零元整"
    integer_part = int(amt)
    decimal_part = round(amt - integer_part, 2)
    if integer_part == 0 and decimal_part == 0:
        return "零元整"
    result = ""
    # Convert integer part
    if integer_part > 0:
        s = str(integer_part)
        groups = []
        while len(s) > 4:
            groups.append(s[-4:]); s = s[:-4]
        groups.append(s)
        for gi, g in enumerate(reversed(groups)):
            bu = big_units[len(groups) - 1 - gi] if gi < len(groups) else ""
            g = g.zfill(4)
            zf = False
            seen = False
            for di in range(4):
                d = int(g[di])
                if d == 0:
                    if seen: zf = True
                else:
                    if zf: result += "零"; zf = False
                    result += digits[d] + units[di]
                    seen = True
            if bu and result and result[-1] != "零":
                result += bu
            elif bu and result:
                result = result.rstrip("零") + bu
        result += "元"
    else:
        result = "零元"
    # Decimal part
    dec_int = int(round(decimal_part * 100))
    if dec_int > 0:
        jiao = dec_int // 10
        fen = dec_int % 10
        if jiao > 0:
            result += digits[jiao] + "角"
        if fen > 0:
            result += digits[fen] + "分"
        if jiao == 0 and fen > 0:
            result += "零" + digits[fen] + "分"
    else:
        result += "整"
    return result

def _merge_carton_groups(ws, products, start_row, end_row, columns):
    groups = {}
    for i, prod in enumerate(products):
        gid = prod.get("carton_group", 0)
        if gid:
            row_num = start_row + i
            if gid not in groups:
                groups[gid] = {"start": row_num, "end": row_num}
            else:
                groups[gid]["end"] = row_num
    for gid, rng in groups.items():
        if rng["start"] == rng["end"]:
            continue
        for col in columns:
            ws.merge_cells(f"{get_column_letter(col)}{rng['start']}:{get_column_letter(col)}{rng['end']}")


def generate_factory_excel(data: dict, image_paths: dict = None) -> bytes:
    if image_paths is None:
        image_paths = {}

    wb = Workbook()
    ws = wb.active
    ws.title = "成品合同"

    # ── Styles ──
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                         top=Side(style="thin"), bottom=Side(style="thin"))
    title_font = Font(name="SimHei", size=20, bold=True)
    normal_font = Font(name="SimSun", size=12)
    small_font = Font(name="SimSun", size=11)
    header_font = Font(name="SimSun", size=11, bold=True)
    clause_font = Font(name="SimSun", size=11)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    company = data.get("company", {})
    factory = data.get("factory", {})
    products = data.get("products", [])
    clauses = data.get("clauses", {})

    # Column widths
    # Left box(A-D) & Right box(F-I) have equal total widths for fair display
    # Left 5 cols (A-E) and Right 5 cols (F-J) have equal total width (56 each)
    for cl, w in [("A", 15), ("B", 12), ("C", 8), ("D", 10), ("E", 11), ("F", 11),
                  ("G", 8), ("H", 12), ("I", 13), ("J", 12)]:
        ws.column_dimensions[cl].width = w

    row = 1

    # ── Title ──
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    ws.cell(row=row, column=1, value="成  品  定  做  合  同").font = title_font
    ws.cell(row=row, column=1).alignment = center_align
    row += 2

    # ── Left-right info (no borders) ──
    info_data = [
        (f"承揽方：{factory.get('name', '')}", f"合同编号：{data.get('contract_no', '')}"),
        (f"定作方：{company.get('name', '')}", f"签订地点：{data.get('sign_place', '')}"),
        ("一、产品名称、款号、数量、金额、交货时间及数量", f"签订时间：{data.get('date', '')}"),
    ]
    for left_val, right_val in info_data:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=10)
        ws.cell(row=row, column=1, value=left_val).font = normal_font
        ws.cell(row=row, column=1).alignment = left_align
        ws.cell(row=row, column=6, value=right_val).font = normal_font
        ws.cell(row=row, column=6).alignment = left_align
        row += 1
    row += 1

    # ── Product table header ──
    headers = ["定作物品或项目", "货号", "颜色", "图片", "数量(PCS)", "装箱数", "箱数", "成品单价(元)", "金额(元)", "交货数量及期限"]
    for i, h in enumerate(headers):
        cell = ws.cell(row=row, column=i + 1, value=h)
        cell.font = header_font; cell.alignment = center_align; cell.border = thin_border
    row += 1

    # ── Product rows ──
    product_start_row = row
    for idx, prod in enumerate(products):
        if prod.get("carton_group_first", True):
            ppc = prod.get("pcs_per_carton", "")
            ctn = prod.get("cartons", "")
        else:
            ppc = ""
            ctn = ""

        ws.cell(row=row, column=1, value=prod.get("name", "")).font = small_font
        ws.cell(row=row, column=1).alignment = left_align; ws.cell(row=row, column=1).border = thin_border

        ws.cell(row=row, column=2, value=prod.get("item_no", "")).font = small_font
        ws.cell(row=row, column=2).alignment = center_align; ws.cell(row=row, column=2).border = thin_border

        ws.cell(row=row, column=3, value=prod.get("color", "")).font = small_font
        ws.cell(row=row, column=3).alignment = center_align; ws.cell(row=row, column=3).border = thin_border

        # Col D: image
        img_key = str(idx)
        if img_key in image_paths and os.path.exists(image_paths[img_key]):
            try:
                img = OpxImage(image_paths[img_key]); img.width = 60; img.height = 60
                ws.add_image(img, f"D{row}")
                ws.row_dimensions[row].height = 70
            except Exception:
                pass
        ws.cell(row=row, column=4).border = thin_border

        ws.cell(row=row, column=5, value=prod.get("qty", "")).font = small_font
        ws.cell(row=row, column=5).alignment = center_align; ws.cell(row=row, column=5).border = thin_border

        ws.cell(row=row, column=6, value=ppc).font = small_font
        ws.cell(row=row, column=6).alignment = center_align; ws.cell(row=row, column=6).border = thin_border

        ws.cell(row=row, column=7, value=ctn).font = small_font
        ws.cell(row=row, column=7).alignment = center_align; ws.cell(row=row, column=7).border = thin_border

        ws.cell(row=row, column=8, value=prod.get("unit_price_cny", "")).font = small_font
        ws.cell(row=row, column=8).alignment = center_align; ws.cell(row=row, column=8).border = thin_border

        ws.cell(row=row, column=9, value=prod.get("amount_cny", "")).font = small_font
        ws.cell(row=row, column=9).alignment = center_align; ws.cell(row=row, column=9).border = thin_border
        ws.cell(row=row, column=10).border = thin_border

        ws.row_dimensions[row].height = max(ws.row_dimensions[row].height or 22, 22)
        row += 1

    _merge_carton_groups(ws, products, product_start_row, row - 1, [6, 7])

    # Merge delivery note column (J) across all product rows
    if products:
        dn = data.get("delivery_note", "20XX年X月X日前全部运送到指定地点")
        ws.merge_cells(start_row=product_start_row, start_column=10, end_row=row - 1, end_column=10)
        ws.cell(row=product_start_row, column=10, value=dn).font = small_font
        ws.cell(row=product_start_row, column=10).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ── TOTAL ──
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.cell(row=row, column=1, value="合 计").font = Font(name="SimSun", size=10, bold=True)
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="right", vertical="center")
    ws.cell(row=row, column=1).border = thin_border
    for ci in [2, 3, 4]:
        ws.cell(row=row, column=ci).border = thin_border
    ws.cell(row=row, column=5, value=data.get("total_qty", "")).font = Font(name="SimSun", size=10, bold=True)
    ws.cell(row=row, column=5).alignment = center_align; ws.cell(row=row, column=5).border = thin_border
    ws.cell(row=row, column=7, value=data.get("total_cartons", "")).font = Font(name="SimSun", size=10, bold=True)
    ws.cell(row=row, column=7).alignment = center_align; ws.cell(row=row, column=7).border = thin_border
    ws.cell(row=row, column=9, value=data.get("total_amount_cny", "")).font = Font(name="SimSun", size=10, bold=True)
    ws.cell(row=row, column=9).alignment = center_align; ws.cell(row=row, column=9).border = thin_border
    row += 1

    # Price note
    pn = data.get("price_note", "成品单价含13%增值税发票，包装，运费")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    ws.cell(row=row, column=1, value=pn).font = Font(name="SimSun", size=9)
    ws.cell(row=row, column=1).alignment = center_align
    row += 1

    # Chinese uppercase total
    cap = num2cn(data.get("total_amount_cny", "0"))
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    ws.cell(row=row, column=1, value=f"合计人民币金额(大写):  人民币{cap}").font = Font(name="SimSun", size=9, bold=True)
    ws.cell(row=row, column=1).alignment = left_align
    row += 2

    # ── Clauses ──
    clause_items = [("二", "clause_1_product"), ("三", "clause_2_quality"), ("四", "clause_3_mold_fee"),
                    ("五", "clause_4_packaging"), ("六", "clause_5_delivery"), ("七", "clause_6_inspection"),
                    ("八", "clause_7_payment"), ("九", "clause_8_guarantee"), ("十", "clause_9_liability"),
                    ("十一", "clause_10_dispute"), ("十二", "clause_11_shipping"), ("十三", "clause_12_other")]
    sep_style = Side(style="thin", color="000000")
    for num, key in clause_items:
        text = clauses.get(key, "")
        display = f"{num}、{text}"
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
        cell = ws.cell(row=row, column=1, value=display)
        cell.font = clause_font
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        h = max(20, len(text) // 60 * 15 + 20)
        ws.row_dimensions[row].height = h
        for ci in range(1, 11):
            ws.cell(row=row, column=ci).border = Border(bottom=sep_style)
        row += 1
    row += 1

    # ── Signatures (boxed: cols 1-4 and 6-9, no border on col 5) ──
    sig_font = Font(name="SimSun", size=10)
    sig_rows_start = row
    sig_data = [
        (f"承揽方(盖章): {factory.get('name', '')}", f"定作方(盖章): {company.get('name', '')}"),
        (f"地址: {factory.get('address', '')}", f"地址: {company.get('address', '')}"),
        (f"税号: {factory.get('tax_id', '')}", f"税号: {company.get('tax_id', '')}"),
    ]
    if factory.get("bank_name"):
        sig_data.append((f"开户行: {factory.get('bank_name', '')}", ""))
        if factory.get("bank_account"):
            sig_data.append((f"账号: {factory.get('bank_account', '')}", ""))

    # Signature: left box A-E, right box F-J, equal width (56 each), no gap
    # Col E has no right border, col F left border is the dividing line
    sig_nowrap = Alignment(horizontal="left", vertical="center", wrap_text=False)
    sig_start = row
    for idx, (lv, rv) in enumerate(sig_data):
        ws.cell(row=row, column=1, value=lv)
        c1f = sig_font if (lv and ("承揽方" in str(lv))) else clause_font
        ws.cell(row=row, column=1).font = c1f; ws.cell(row=row, column=1).alignment = sig_nowrap
        ws.cell(row=row, column=6, value=rv)
        c6f = sig_font if (rv and ("定作方" in str(rv))) else clause_font
        ws.cell(row=row, column=6).font = c6f; ws.cell(row=row, column=6).alignment = sig_nowrap
        row += 1
    sig_end = row - 1

    thin = Side(style="thin", color="000000")
    # Left box A-E: only outer edges (no right border on E — F's left edge is divider)
    for r in range(sig_start, sig_end + 1):
        for c in range(1, 6):
            ws.cell(row=r, column=c).border = Border(
                left=thin if c == 1 else None,
                right=None,  # no right border on left box
                top=thin if r == sig_start else None,
                bottom=thin if r == sig_end else None)
    # Right box F-J: left border on F = divider, outer edges on J
    for r in range(sig_start, sig_end + 1):
        for c in range(6, 11):
            ws.cell(row=r, column=c).border = Border(
                left=thin if c == 6 else None,  # col 6 left = visible divider
                right=thin if c == 10 else None,
                top=thin if r == sig_start else None,
                bottom=thin if r == sig_end else None)

    # ── Print settings ──
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = 9
    ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_margins.left = 0.3; ws.page_margins.right = 0.3
    ws.page_margins.top = 0.4; ws.page_margins.bottom = 0.4

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
